# --- Carregar as bibliotecas necessárias ---
import pandas as pd
import numpy as np
import requests
import openpyxl
import re
import os

print("✅ Bibliotecas carregadas")

# ===================================================================
# ---- CONFIGURAÇÕES ----
# ===================================================================
# Caminhos ancorados no diretório do script (e não no cwd), para funcionar
# tanto quando é corrido a partir da raiz do repositório como de scripts/.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(SCRIPT_DIR)

FICHEIRO_EXCEL = os.path.join(ROOT_DIR, "data", "simuladores",
                              "simulador-autoconsumo",
                              "autoconsumo_Tiago_Felicia.xlsx")
ABA_EXCEL = "OMIE_CICLOS"
COLUNA_PARA_ESCREVER = 8 # Coluna H
# Intermédio da Fase 1 (atualizar_mibel_ano_atual_ACUM.py) — partilhado
FICHEIRO_MIBEL_CSV = os.path.join(ROOT_DIR, "data", "omie", "MIBEL_ano_atual_ACUM.csv")

print(f"ℹ️ Fonte de dados: '{FICHEIRO_MIBEL_CSV}'")
print("⚠️ MODO: Apenas dados reais (sem futuros)")
# ===================================================================

def run_update_process():
    """
    Função principal que encapsula todo o processo de ETL (SEM FUTUROS).
    """
    try:
        # ========================================================
        # PASSO 1: Leitura dos Dados
        # ========================================================
        
        print(f"\n⏳ Passo 1: A ler dados históricos do '{FICHEIRO_MIBEL_CSV}'...")
        try:
            dados_combinados_qh = pd.read_csv(FICHEIRO_MIBEL_CSV, parse_dates=['Data'])
            
            # Este script usa internamente a coluna 'Preco' para o preço de PT
            dados_combinados_qh = dados_combinados_qh.rename(columns={'Preco_PT': 'Preco'})
            
            # Selecionar apenas as colunas que precisamos
            # Nota: Este script só usa dados REAIS, por isso filtramos logo por Preco notna()
            dados_combinados_qh = dados_combinados_qh[['Data', 'Hora', 'Preco']]
            dados_combinados_qh = dados_combinados_qh.dropna(subset=['Data', 'Hora', 'Preco'])
            
            ultima_data_omie = dados_combinados_qh['Data'].max()
            print(f"✅ {len(dados_combinados_qh)} registos históricos (com dados) lidos.")
            print(f"   📅 Última data com dados: {ultima_data_omie.date()}")
            
        except FileNotFoundError:
            print(f"❌ ERRO CRÍTICO: O ficheiro '{FICHEIRO_MIBEL_CSV}' não foi encontrado.")
            print("   - Por favor, execute primeiro o script 'update_mibel_historico.py'.")
            return
        except Exception as e:
            print(f"❌ ERRO CRÍTICO ao ler o ficheiro histórico: {e}")
            return

        # =================================================================
        # PASSO 2: Criar estrutura completa e converter para Portugal
        # =================================================================

        print("\n⏳ Passo 2: A criar estrutura completa até 2026...")
        
        def num_quartos_dia(data_obj):
            """
            Calcula número de quartos horários considerando DST.
            Usa a diferença entre dias para garantir que apanha 23h ou 25h.
            """
            tz_es = 'Europe/Madrid'
            
            # Garantir que estamos a usar apenas a data
            dia_atual = data_obj.date() if hasattr(data_obj, 'date') else data_obj
            dia_seguinte = dia_atual + pd.Timedelta(days=1)
            
            # Criar Timestamps localizados (00:00 hoje vs 00:00 amanhã)
            dt0 = pd.Timestamp(f"{dia_atual} 00:00:00", tz=tz_es)
            dt_next = pd.Timestamp(f"{dia_seguinte} 00:00:00", tz=tz_es)
            
            horas = (dt_next - dt0).total_seconds() / 3600
            return int(round(horas * 4)) 
        
        datas_futuras = pd.date_range(start=ultima_data_omie + pd.Timedelta(days=1), end='2026-12-31', freq='D')
        
        futuro_qh = []
        for data in datas_futuras:
            n_quartos = num_quartos_dia(data)
            for hora in range(1, n_quartos + 1):
                futuro_qh.append({'Data': data, 'Hora': hora, 'Preco': np.nan})
        
        if futuro_qh:
            futuro_qh = pd.DataFrame(futuro_qh)
        else:
            futuro_qh = pd.DataFrame(columns=['Data', 'Hora', 'Preco'])

        # Combinar histórico + estrutura futura (vazia)
        dados_completos_qh = pd.concat([dados_combinados_qh, futuro_qh], ignore_index=True)
        dados_completos_qh = dados_completos_qh.sort_values(['Data', 'Hora']).reset_index(drop=True)
        
        print(f"   - Dados reais até: {ultima_data_omie.date()}")
        print(f"   - Estrutura criada até: 2026-12-31")
        print(f"   - Registos com dados reais: {len(dados_combinados_qh)}")
        print(f"   - Registos vazios (futuros): {len(futuro_qh)}")
        
        # ============================================================
        # PASSO 3: Conversão para hora de Portugal
        # ============================================================
        print("\n⏳ Passo 3: A converter para hora de Portugal...")
        
        # Gerar datetime em hora de Espanha
        def gerar_datetime_es(row):
            """Gera timestamp correto considerando DST"""
            data = row['Data']
            hora = row['Hora']
            inicio_dia = pd.Timestamp(f"{data} 00:00:00", tz='Europe/Madrid')
            return inicio_dia + pd.Timedelta(minutes=15 * (hora - 1))
        
        dados_completos_qh['datetime_es'] = dados_completos_qh.apply(gerar_datetime_es, axis=1)
        dados_completos_qh['datetime_pt'] = dados_completos_qh['datetime_es'].dt.tz_convert('Europe/Lisbon')
        dados_completos_qh['Data_PT'] = dados_completos_qh['datetime_pt'].dt.date
        
        # Renumerar horas em hora de Portugal
        dados_finais_pt = dados_completos_qh.sort_values('datetime_pt').copy()
        dados_finais_pt['Hora_PT'] = dados_finais_pt.groupby('Data_PT').cumcount() + 1
        
        # Selecionar apenas 2026 e 2027
        dados_finais_pt = dados_finais_pt[dados_finais_pt['datetime_pt'].dt.year.isin([2026, 2027])].copy()
        dados_finais_pt = dados_finais_pt[['Data_PT', 'Hora_PT', 'Preco']].rename(
            columns={'Data_PT': 'Data', 'Hora_PT': 'Hora'}
        )
        dados_finais_pt = dados_finais_pt.reset_index(drop=True)
        
        registos_com_dados = dados_finais_pt['Preco'].notna().sum()
        registos_vazios = dados_finais_pt['Preco'].isna().sum()
        
        print(f"✅ {len(dados_finais_pt)} registos totais preparados em hora de Portugal.")
        print(f"   - Com dados reais: {registos_com_dados}")
        print(f"   - Vazios (NaN): {registos_vazios}")
        
        # Validação de quartos
        check_quartos = dados_finais_pt.groupby('Data').size().reset_index(name='n')
        dias_estranhos = check_quartos[~check_quartos['n'].isin([92, 96, 100])]
        
        if not dias_estranhos.empty:
            print("⚠️ Aviso: Dias com número de quartos inesperado:")
            print(dias_estranhos.to_string(index=False))
        else:
            print("✅ Todos os dias têm número de quartos esperado (92, 96 ou 100).")

        # ============================================================
        # PASSO 4: Atualização do ficheiro Excel
        # ============================================================
        print(f"\n⏳ Passo 4: A atualizar o ficheiro '{FICHEIRO_EXCEL}'...")

        # 1. Ler a pauta de tempo 'master' do Excel (Colunas A e B)
        print(f"   - A ler a pauta de tempo da aba '{ABA_EXCEL}' para alinhamento...")
        df_pauta_excel = pd.read_excel(
            FICHEIRO_EXCEL,
            sheet_name=ABA_EXCEL,
            usecols=['Data', 'Hora'] 
        )
        df_pauta_excel.dropna(subset=['Data', 'Hora'], inplace=True)
        # Preservar a ordem original do Excel (o seu índice 0-based)
        df_pauta_excel = df_pauta_excel.reset_index() 
        
        # 2. Preparar a pauta do Excel para o merge
        df_pauta_excel['Data'] = pd.to_datetime(df_pauta_excel['Data']).dt.date
        # Gerar a sequência numérica (1-96, 1-100, etc.)
        df_pauta_excel['Hora'] = df_pauta_excel.groupby('Data').cumcount() + 1
        
        # 3. Preparar os nossos dados calculados (do Passo 3)
        df_dados_pt_merge = dados_finais_pt.copy()
        df_dados_pt_merge['Data'] = pd.to_datetime(df_dados_pt_merge['Data']).dt.date
        df_dados_pt_merge['Hora'] = df_dados_pt_merge['Hora'].astype(int)
        
        # ===================================================================
        # Adicionar um marcador para saber quais linhas são de 2026/2027
        df_dados_pt_merge['dados_calculados'] = True
        # ===================================================================

        # 4. Fazer o MERGE para alinhar os preços à pauta do Excel
        print("   - A alinhar preços calculados com a pauta do Excel...")
        df_final_excel = pd.merge(
            df_pauta_excel,
            df_dados_pt_merge[['Data', 'Hora', 'Preco', 'dados_calculados']], # Trazer o marcador
            on=['Data', 'Hora'],
            how='left' # Manter todas as linhas da pauta
        )
        
        # 5. Ordenar pela ordem original do Excel
        df_final_excel = df_final_excel.sort_values('index').reset_index(drop=True)
        
        # 6. FILTRAR apenas os dados que TÊM o marcador 'dados_calculados'
        dados_para_escrever = df_final_excel[df_final_excel['dados_calculados'] == True].copy()
        
        print(f"   - {len(dados_para_escrever)} preços (2026/2027) alinhados e prontos a escrever.")

        # 7. Escrever no ficheiro Excel (de forma seletiva)
        print(f"   - A carregar '{FICHEIRO_EXCEL}' para escrita...")
        wb = openpyxl.load_workbook(FICHEIRO_EXCEL)
        sheet = wb[ABA_EXCEL]
        
        print(f"   - A escrever {len(dados_para_escrever)} preços na Coluna {COLUNA_PARA_ESCREVER} (H)...")
        
        # Iterar APENAS sobre as linhas que TÊM dados
        for _, row in dados_para_escrever.iterrows():
            # Usar o 'index' original para encontrar a linha correta no Excel
            excel_row_index = int(row['index']) + 2  # +1 (0-based to 1-based) +1 (skip header)
            preco = row['Preco'] # Este 'preco' PODE ser NaN (se for futuro)
            
            if pd.isna(preco):
                sheet.cell(row=excel_row_index, column=COLUNA_PARA_ESCREVER, value=None)
            else:
                sheet.cell(row=excel_row_index, column=COLUNA_PARA_ESCREVER, value=preco)
            
        # ===================================================================
            
        # 8. Atualizar data de referência na aba 'Constantes'
        sheet_const = wb["Constantes"]
        sheet_const['B42'] = ultima_data_omie.strftime('%d/%m/%Y')
        
        wb.save(FICHEIRO_EXCEL)
        print(f"✅ O ficheiro Excel foi atualizado com sucesso!")
        print(f"   Data_Valores_OMIE = {ultima_data_omie.date()}")
        print(f"   ⚠️ Nota: Apenas dados reais até {ultima_data_omie.date()} foram escritos.")

    except Exception as e:
        import traceback
        print(f"❌ Ocorreu um erro inesperado: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    run_update_process()